import tkinter as tk
from tkinter import scrolledtext
from tkinter import ttk
import openpyxl

class ExcelChatbot:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.questions = []
        self.categorias_mapeadas = {
            "esperanza de vida hombre": "esperanza de vida al nacer (años) (hombre)",
            "esperanza de vida mujer": "esperanza de vida al nacer (años) (mujer)",
            "composición 0 a 14": "composición de la población en años (porcentaje) (0 a 14)",
            "composición 15 a 64": "composición de la población en años (porcentaje) (15 a 64)",
            "composición 65 a más": "composición de la población en años (porcentaje) (65 a más)",
            "fecundidad total": "tasa de fecundidad total",
            "poblacion": "población"
        }

    def combinar_encabezados(self, sheet):
        encabezados = []

        for col in range(1, sheet.max_column + 1):
            encabezado_actual = ""
            valor_fila_1 = sheet.cell(row=1, column=col).value
            valor_fila_2 = sheet.cell(row=2, column=col).value
            valor_fila_3 = sheet.cell(row=3, column=col).value
            valor_fila_4 = sheet.cell(row=4, column=col).value

            if valor_fila_1 and valor_fila_1.strip():
                encabezado_actual = valor_fila_1.strip().lower()
            if valor_fila_2 and valor_fila_2.strip():
                encabezado_actual += f" ({valor_fila_2.strip().lower()})"
            if valor_fila_3 and valor_fila_3.strip():
                encabezado_actual += f" ({valor_fila_3.strip().lower()})"
            if valor_fila_4 and valor_fila_4.strip():
                encabezado_actual += f" ({valor_fila_4.strip().lower()})"

            encabezados.append(encabezado_actual)

        return [encabezado.strip() for encabezado in encabezados]

    def obtener_valor(self, pais, categoria):
        wb = openpyxl.load_workbook(self.excel_file)
        sheet = wb.active

        encabezados = self.combinar_encabezados(sheet)
        encabezados = [encabezado.strip().lower() for encabezado in encabezados]

        for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
            if row[0] is not None:
                nombre_pais = row[0].strip().lower()
                if pais.lower() == nombre_pais:
                    categoria_mapeada = self.categorias_mapeadas.get(categoria)
                    if categoria_mapeada in encabezados:
                        indice_categoria = encabezados.index(categoria_mapeada)
                        return row[indice_categoria]

        return None

    def obtener_respuesta(self, pregunta, pais, operacion=None, pais2=None):
        if "esperanza de vida al nacer para hombres" in pregunta:
            categoria = "esperanza de vida hombre"
        elif "esperanza de vida al nacer para mujeres" in pregunta:
            categoria = "esperanza de vida mujer"
        elif "composición de la población de 0 a 14 años" in pregunta:
            categoria = "composición 0 a 14"
        elif "composición de la población de 15 a 64 años" in pregunta:
            categoria = "composición 15 a 64"
        elif "composición de la población de 65 a más años" in pregunta:
            categoria = "composición 65 a más"
        elif "tasa de fecundidad total" in pregunta:
            categoria = "fecundidad total"
        elif "población" in pregunta:
            categoria = "poblacion"
        else:
            return "Pregunta no reconocida."

        valor1 = self.obtener_valor(pais, categoria)

        if operacion and pais2:
            valor2 = self.obtener_valor(pais2, categoria)
            if valor1 is not None and valor2 is not None:
                if operacion == "suma":
                    resultado = valor1 + valor2
                elif operacion == "resta":
                    resultado = valor1 - valor2
                elif operacion == "multiplicación":
                    resultado = valor1 * valor2
                elif operacion == "división":
                    if valor2 != 0:
                        resultado = valor1 / valor2
                    else:
                        return "No se puede dividir por cero."
                return f"La {operacion} de la {categoria} entre {pais} y {pais2} es {resultado}."
            else:
                return f"No encontré información para realizar la {operacion} entre {pais} y {pais2}."
        else:
            if valor1 is not None:
                return f"La {categoria} en {pais} es {valor1}."
            else:
                return f"No encontré información sobre {categoria} para {pais}."

    def answer_question(self, pregunta, pais, operacion=None, pais2=None):
        self.questions.append((pregunta, pais, operacion, pais2))
        return self.obtener_respuesta(pregunta, pais, operacion, pais2)

def submit_question(event=None):
    pregunta = question_combo.get()
    pais = country_combo.get()
    operacion = operation_combo.get()
    pais2 = country_combo2.get()

    if pregunta and pais:
        if operacion and pais2:
            chatbot_response = chatbot.answer_question(pregunta, pais, operacion, pais2)
        else:
            chatbot_response = chatbot.answer_question(pregunta, pais)
        chat_log.insert(tk.END, f"Usuario: {pregunta} en {pais}\n")
        if operacion and pais2:
            chat_log.insert(tk.END, f"Operación: {operacion} con {pais2}\n")
        chat_log.insert(tk.END, f"Chatbot: {chatbot_response}\n\n")

def show_questions():
    chat_log.insert(tk.END, "Historial de preguntas:\n")
    for q, p, o, p2 in chatbot.questions:
        if o and p2:
            chat_log.insert(tk.END, f"- {q} entre {p} y {p2} ({o})\n")
        else:
            chat_log.insert(tk.END, f"- {q} en {p}\n")
    chat_log.insert(tk.END, "\n")

def clear_console():
    chat_log.delete(1.0, tk.END)

# Ruta del archivo Excel
file_path = 'D:\Chatbot\Demografia.xlsx'

# Inicializar el chatbot con los datos del Excel
chatbot = ExcelChatbot(file_path)

# Obtener la lista de países
wb = openpyxl.load_workbook(file_path)
sheet = wb.active
countries = [row[0] for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True) if row[0]]

# Crear la interfaz de usuario
root = tk.Tk()
root.title("Asistente de consultas demográficas")

mainframe = ttk.Frame(root, padding="10 10 10 10")
mainframe.grid(column=0, row=0, sticky=(tk.W, tk.E, tk.N, tk.S))

root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

question_label = ttk.Label(mainframe, text="Seleccione su pregunta:")
question_label.grid(column=1, row=1, sticky=(tk.W, tk.E))

questions = [
    "¿Cuál es la población?",
    "¿Cuál es la composición de la población de 0 a 14 años?",
    "¿Cuál es la composición de la población de 15 a 64 años?",
    "¿Cuál es la composición de la población de 65 a más años?",
    "¿Cuál es la tasa de fecundidad total?",
    "¿Cuál es la esperanza de vida al nacer para hombres?",
    "¿Cuál es la esperanza de vida al nacer para mujeres?",
]

question_combo = ttk.Combobox(mainframe, values=questions, width=80)
question_combo.grid(column=1, row=2, sticky=(tk.W, tk.E))

country_label = ttk.Label(mainframe, text="Seleccione el país:")
country_label.grid(column=1, row=3, sticky=(tk.W, tk.E))

country_combo = ttk.Combobox(mainframe, values=countries, width=80)
country_combo.grid(column=1, row=4, sticky=(tk.W, tk.E))

country_label2 = ttk.Label(mainframe, text="Seleccione el segundo país (para operaciones):")
country_label2.grid(column=1, row=5, sticky=(tk.W, tk.E))

country_combo2 = ttk.Combobox(mainframe, values=countries, width=80)
country_combo2.grid(column=1, row=6, sticky=(tk.W, tk.E))

operation_label = ttk.Label(mainframe, text="Seleccione la operación:")
operation_label.grid(column=1, row=7, sticky=(tk.W, tk.E))

operations = ["suma", "resta", "multiplicación", "división"]
operation_combo = ttk.Combobox(mainframe, values=operations, width=80)
operation_combo.grid(column=1, row=8, sticky=(tk.W, tk.E))

submit_button = ttk.Button(mainframe, text="Enviar", command=submit_question)
submit_button.grid(column=2, row=2, sticky=(tk.W, tk.E))

show_questions_button = ttk.Button(mainframe, text="Historial de consultas", command=show_questions)
show_questions_button.grid(column=2, row=3, sticky=(tk.W, tk.E))

clear_button = ttk.Button(mainframe, text="Limpiar consola", command=clear_console)
clear_button.grid(column=2, row=4, sticky=(tk.W, tk.E))

chat_log = scrolledtext.ScrolledText(mainframe, width=100, height=20)
chat_log.grid(column=1, row=9, columnspan=2, sticky=(tk.W, tk.E))

for child in mainframe.winfo_children():
    child.grid_configure(padx=5, pady=5)

root.mainloop()
